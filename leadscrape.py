import sys
import subprocess

def ensure(package):
    try:
        __import__(package)
    except ImportError:
        print(f"📦 Installing missing package: {package}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

ensure("pandas")
ensure("googlemaps")
ensure("openpyxl")
ensure("requests")
ensure("beautifulsoup4")

import pandas as pd
import requests
import re
import time
import os
from bs4 import BeautifulSoup
from googlemaps import Client as GoogleMaps
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side

API_KEY = 'your_key_here'
gmaps = GoogleMaps(API_KEY)

PLACE_DETAILS_COST_PER_REQUEST = 0.017
NEARBY_SEARCH_COST_PER_REQUEST = 0.032

def get_place_details(place_id):
    try:
        place_details = gmaps.place(place_id=place_id, fields=[
            'name',
            'formatted_phone_number',
            'website',
            'rating',
            'user_ratings_total',
            'url'
        ])
        return place_details['result']
    except Exception as e:
        print(f"[ERROR] Failed to fetch details for place_id {place_id}: {e}")
        return {}

def extract_email_from_website(url):
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        res = requests.get(url, timeout=5, headers=headers)
        emails = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", res.text)
        return emails[0] if emails else "N/A"
    except Exception:
        return "N/A"

def autosize_and_style_excel(filename, leads):
    wb = load_workbook(filename)
    ws = wb.active

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    for i, lead in enumerate(leads, start=2):
        maps_url = lead.get('Google Maps URL', '')
        if maps_url and maps_url != 'N/A':
            cell = ws[f"F{i}"]
            cell.value = "View on Maps"
            cell.hyperlink = maps_url
            cell.style = "Hyperlink"

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    ws.auto_filter.ref = ws.dimensions
    wb.save(filename)

def estimate_cost(results_count, pages=1):
    details_cost = results_count * PLACE_DETAILS_COST_PER_REQUEST
    search_cost = pages * NEARBY_SEARCH_COST_PER_REQUEST
    return round(details_cost + search_cost, 4)

def main(industry, area, radius):
    print(f"\n📍 Geocoding location: {area}")
    try:
        geocode_result = gmaps.geocode(area)
        if not geocode_result:
            raise ValueError("No geocode result returned.")
        location = geocode_result[0]['geometry']['location']
        print(f"✅ Location found: {location}")
    except Exception as e:
        import traceback
        print("[ERROR] Geocoding failed:")
        traceback.print_exc()
        return

    ESTIMATED_PAGES = 3
    ESTIMATED_RESULTS = ESTIMATED_PAGES * 20
    cost_estimate = estimate_cost(ESTIMATED_RESULTS, ESTIMATED_PAGES)
    print(f"\n💸 Estimated API Cost: ${cost_estimate} (based on ~{ESTIMATED_RESULTS} businesses)")
    proceed = input("🚦 Press Enter to continue or type 'n' to cancel: ").strip().lower()
    if proceed == 'n':
        print("❌ Cancelled.")
        return

    all_leads = []
    next_page_token = None
    page = 1

    while True:
        print(f"🔍 Searching page {page} for '{industry}'...")
        try:
            if next_page_token:
                for attempt in range(5):
                    time.sleep(2)
                    places = gmaps.places_nearby(
                        location=location,
                        radius=radius,
                        keyword=industry,
                        page_token=next_page_token
                    )
                    if 'results' in places and places['results']:
                        break
                else:
                    print("⚠️  next_page_token didn't return results, breaking.")
                    break
            else:
                places = gmaps.places_nearby(location=location, radius=radius, keyword=industry)
        except Exception as e:
            print(f"[ERROR] Failed to fetch places: {e}")
            break

        results = places.get('results', [])
        print(f"➡️  Found {len(results)} results on page {page}")
        all_leads.extend(results)

        next_page_token = places.get('next_page_token')
        if not next_page_token:
            break

        page += 1

    print(f"📦 Total raw places found: {len(all_leads)}")

    leads = []

    for i, place in enumerate(all_leads, 1):
        print(f"📞 Fetching details for: {place.get('name', 'Unknown')} ({i}/{len(all_leads)})")
        details = get_place_details(place['place_id'])

        website = details.get('website', 'N/A')
        email = extract_email_from_website(website) if website != "N/A" else "N/A"

        lead = {
            'Name': details.get('name', 'N/A'),
            'Phone Number': details.get('formatted_phone_number', 'N/A'),
            'Website': website,
            'Rating': details.get('rating', 'N/A'),
            'Total Reviews': details.get('user_ratings_total', 'N/A'),
            'Google Maps URL': details.get('url', 'N/A'),
            'Email Address': email
        }
        leads.append(lead)

    print(f"✅ Total leads exported: {len(leads)}")

    filename = f"{industry.replace(' ', '_')}_in_{area.replace(',', '').replace(' ', '_')}.xlsx"
    df = pd.DataFrame(leads)
    df.to_excel(filename, index=False)
    autosize_and_style_excel(filename, leads)

    print(f"📄 Saved to: {filename}")
    input("\n🟢 Done! Press Enter to open the file...")
    os.startfile(filename)

if __name__ == '__main__':
    print("🧠 Lead Scraper")
    industry = input("👉 Enter industry to search for: ").strip()
    area = input("🌍 Enter area to search in (e.g. London, UK): ").strip()
    radius_input = input("📏 Enter search radius in metres (default is 5000): ").strip()
    radius = int(radius_input) if radius_input.isdigit() else 5000
    main(industry, area, radius)
